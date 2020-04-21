# coding: utf-8

import codecs  # 文字コーデック指定
import re  # 正規表現での区切り文字指定
import openpyxl  # エクセル操作


class Converter():
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.nameList = []  # 測定データのタイトルリスト
        self.dataList = []  # 測定データ
        self.summaryList = []  # 各測定データの集計値
        self.summaryIndex = []  #
        self.srslist = []  # 各測定データシート中のグラフリスト。summaryシートに載せる

    def convert(self, input_filename, excelfilename):
        self.wb = openpyxl.Workbook()
        self.nameList = []  # 測定データのタイトルリスト
        self.dataList = []  # 測定データ
        self.summaryList = []  # 各測定データの集計値
        self.summaryIndex = []  #
        self.srslist = []  # 各測定データシート中のグラフリスト。summaryシートに載せる
        self.read_input_file(input_filename)
        self.create_sheet()
        self.write_each_sheet()
        self.write_summary_sheet()
        self.save_excel(excelfilename)

    def read_input_file(self, input_filename):  # テキストファイル読み込み
        iPoint = 0
        count = 0
        with codecs.open(inputfilename, "r", "cp932") as input_file:
            for line in input_file:
                if line == "" or line == "\n" or line == "\r\n":
                    continue
                elif count == 0:
                    itemList = re.split('No.:|;Name:|;NoOfPnt:|\r\n', line)
                    Name = str.strip(itemList[2])
                    num = int(itemList[3])
                    count += 1
                    temp = [[0, 0, 0] for i in range(num)]
                    self.nameList.append(Name)
                else:
                    if count < num + 1:
                        itemList = re.split('X:|;Y:|;Z:|;D:|;d:|\r\n', line)
                        x = float(itemList[1])
                        y = float(itemList[2])
                        z = float(itemList[3])
                        temp[count - 1] = [x, y, z]
                        count += 1
                    elif count == num + 1:
                        self.dataList.append(temp)
                        itemList = re.split(';|:|\r\n', line)
                        temp_s = []
                        for s in itemList[1:-1:2]:
                            try:
                                temp_s.append(float(s))
                            except:
                                temp_s.append('')

                        self.summaryList.append(temp_s)
                        self.summaryIndex.append(itemList[0:-1:2])

                        count = 0
                        iPoint += 1

            # 名前重複処理(AAA, AAA_2, AAA_3...とする)
            newNameList = ["" for i in range(len(self.nameList))]
            for i, iname in enumerate(self.nameList):
                addNumber = self.nameList[0:i].count(iname)
                if addNumber > 0:
                    newNameList[i] = self.nameList[i] + '_' + str(addNumber)
                else:
                    newNameList[i] = self.nameList[i]
            self.nameList = newNameList

    def create_sheet(self):
        # 各データの名前でシート作成
        self.sheet = self.wb.active
        self.sheet.title = self.nameList[0]
        for i in range(1, len(self.nameList)):
            self.wb.create_sheet(title=self.nameList[i])
        # summary sheet作成
        self.wb.create_sheet(index=0, title='summary')

    def write_each_sheet(self):
        for i in range(len(self.nameList)):
            ws = self.wb[self.nameList[i]]
            count = 0
            ws.cell(row=1, column=1, value='Name')
            ws.cell(row=1, column=2, value=self.nameList[i])
            for j in range(len(self.summaryIndex[i])):
                ws.cell(row=count + 2, column=1, value=self.summaryIndex[i][count])
                ws.cell(row=count + 2, column=2, value=self.summaryList[i][j])
                count += 1

            ws.cell(row=8, column=1, value='NoOfPnt')
            noOfPnt = len(self.dataList[i])
            ws.cell(row=8, column=2, value=noOfPnt)
            ws.cell(row=9, column=1, value='X')
            ws.cell(row=9, column=2, value='Y')
            ws.cell(row=9, column=3, value='Z')
            ws.cell(row=9, column=4, value='theta')
            ws.cell(row=9, column=5, value='r')
            ws.cell(row=9, column=7, value='X_difr')
            ws.cell(row=9, column=8, value='Y_difr')
            ws.cell(row=9, column=9, value='Include')

            for j in range(noOfPnt):
                ws.cell(row=j + 10, column=1, value=self.dataList[i][j][0])
                ws.cell(row=j + 10, column=2, value=self.dataList[i][j][1])
                ws.cell(row=j + 10, column=3, value=self.dataList[i][j][2])
                calc_theta = '=ATAN2(A' + str(j + 10) + '-$L$2,B' + str(j + 10) + '-$L$3)'
                calc_r = '=((A' + str(j + 10) + '-$L$2)^2+(B' + str(j + 10) + '-$L$3)^2)^0.5'
                ws.cell(row=j + 10, column=4, value=calc_theta)
                ws.cell(row=j + 10, column=5, value=calc_r)
                ws.cell(row=j + 10, column=7, value='=IF(I' + str(j + 10) + ',(E' + str(j + 10) + '-$H$6)*cos(D' + str(
                    j + 10) + '+$H$7)*1000,NA())')
                ws.cell(row=j + 10, column=8, value='=IF(I' + str(j + 10) + ',(E' + str(j + 10) + '-$H$6)*sin(D' + str(
                    j + 10) + '+$H$7)*1000,NA())')
                ws.cell(row=j + 10, column=9, value='=TRUE')

            ws.cell(row=noOfPnt + 10, column=7, value='=G10')
            ws.cell(row=noOfPnt + 10, column=8, value='=H10')

            ws.cell(row=4, column=7, value='r_max')
            ws.cell(row=4, column=8, value='=max(E10:E' + str(noOfPnt + 9) + ')')
            ws.cell(row=5, column=7, value='r_min')
            ws.cell(row=5, column=8, value='=min(E10:E' + str(noOfPnt + 9) + ')')
            ws.cell(row=6, column=7, value='r_0')
            ws.cell(row=6, column=8, value="='summary'!I" + str(i + 2))
            ws.cell(row=7, column=7, value="theta_rot")
            ws.cell(row=7, column=8, value="='summary'!J" + str(i + 2))

            x_range = 'A10:A' + str(noOfPnt + 9)
            y_range = 'B10:B' + str(noOfPnt + 9)
            boo_range = 'I10:I' + str(noOfPnt + 9)

            # 点の中心座標、直径、真円度を計算 (ノイズ除去を想定した形式)
            ws['J1'] = '再計算'
            ws['K1'] = 'Name'
            ws['L1'] = '=B1'
            ws['K2'] = 'X'
            ws['L2'] = '=-P13/2'
            ws['K3'] = 'Y'
            ws['L3'] = '=-P14/2'
            ws['K4'] = 'Z'
            ws['L4'] = '=B4'
            ws['K5'] = 'D'
            ws['L5'] = '=SQRT(L2^2+L3^2-P15)*2'
            ws['K6'] = 'd'
            temp_str = '(IF(' + boo_range + ',SQRT((' + x_range + '-L2)^2+(' + y_range + '-L3)^2),""))'
            ws['L6'] = '=MAX' + temp_str + '-MIN' + temp_str
            ws.formula_attributes['L6'] = {'t': 'array', 'ref': "L6"}

            ws['K9'] = 'N'
            ws['L9'] = '=COUNTIF(' + boo_range + ',TRUE)'
            ws['K10'] = 'X'
            ws['L10'] = '=SUMPRODUCT(' + x_range + '*' + boo_range + ')'
            ws['K11'] = 'Y'
            ws['L11'] = '=SUMPRODUCT(' + y_range + '*' + boo_range + ')'
            ws['K12'] = 'X2'
            ws['L12'] = '=SUMPRODUCT(' + x_range + '^2*' + boo_range + ')'
            ws['K13'] = 'Y2'
            ws['L13'] = '=SUMPRODUCT(' + y_range + '^2*' + boo_range + ')'
            ws['K14'] = 'XY'
            ws['L14'] = '=SUMPRODUCT(' + x_range + '*' + y_range + '*' + boo_range + ')'
            ws['K15'] = 'X3+XY2'
            ws['L15'] = '=SUMPRODUCT(' + boo_range + '*(' + x_range + '^3+' + x_range + '*' + y_range + '^2))'
            ws['K16'] = 'Y3+YX2'
            ws['L16'] = '=SUMPRODUCT(' + boo_range + '*(' + y_range + '^3+' + y_range + '*' + x_range + '^2))'
            ws['K17'] = 'X2+Y2'
            ws['L17'] = '=L12+L13'

            ws['O9'] = '=L12'
            ws['O10'] = '=L14'
            ws['O11'] = '=L10'
            ws['P9'] = '=L14'
            ws['P10'] = '=L13'
            ws['P11'] = '=L11'
            ws['Q9'] = '=L10'
            ws['Q10'] = '=L11'
            ws['Q11'] = '=L9'

            ws['O13'] = 'A'
            ws['O14'] = 'B'
            ws['O15'] = 'C'
            ws['P13'] = '=MMULT(MINVERSE(O9:Q11),-L15:L17)'
            ws.formula_attributes['P13'] = {'t': 'array', 'ref': 'P13:p15'}

            temp_r = [((self.dataList[i][j][0] - self.summaryList[i][0]) ** 2 +
                       (self.dataList[i][j][1] - self.summaryList[i][1]) ** 2) ** 0.5
                      for j in range(noOfPnt)]

            # グラフの縦横軸最大値を設定(正方形グラフを作りたいので1変数)
            scale = round((max(temp_r) - (min(temp_r) - 0.005)) * 1000)
            scale += (5 - scale % 5) % 5

            # draw graph
            cht = openpyxl.chart.ScatterChart()
            cp = openpyxl.drawing.text.CharacterProperties(sz=1000)
            xvalues = openpyxl.chart.Reference(ws, min_col=7, min_row=10,
                                               max_col=7, max_row=noOfPnt + 10)
            yvalues = openpyxl.chart.Reference(ws, min_col=8, min_row=9,
                                               max_col=8, max_row=noOfPnt + 10)
            title = openpyxl.chart.Reference(ws, min_col=2, min_row=1,
                                             max_col=2, max_row=1)

            srs = openpyxl.chart.Series(values=yvalues, xvalues=xvalues, title_from_data=True)
            srs.graphicalProperties.line.dashStyle = 'sysDot'  # memo Value must be one of {'sysDash', 'lgDash', 'sysDashDot', 'lgDashDot', 'sysDot', 'solid', 'sysDashDotDot', 'dash', 'lgDashDotDot', 'dashDot', 'dot'}

            srs.title.strRef.f = str(title)

            srs.marker.symbol = 'circle'
            self.srslist.append(srs)
            cht.series.append(srs)
            cht.width = 15
            cht.height = 15
            cht.layout = openpyxl.chart.layout.Layout(manualLayout=
            openpyxl.chart.layout.ManualLayout(
                layoutTarget='inner',
                h=0.8, w=0.8,
                xMode="edge", yMode="edge"))
            cht.x_axis.scaling.min = -scale
            cht.x_axis.scaling.max = scale
            cht.y_axis.scaling.min = -scale
            cht.y_axis.scaling.max = scale
            cht.x_axis.majorUnit = scale / 5
            cht.y_axis.majorUnit = scale / 5
            cht.legend = None
            ws.add_chart(cht, 'L10')

    def write_summary_sheet(self):
        ws = self.wb['summary']
        ws.cell(row=1, column=1, value='Name')
        count = 0
        for i in range(len(self.summaryIndex[0])):
            ws.cell(row=1, column=count + 2, value=self.summaryIndex[0][count])
            count += 1
        ws.cell(row=1, column=count + 2, value="r_max")
        ws.cell(row=1, column=count + 3, value="r_min")
        ws.cell(row=1, column=count + 4, value="r_0")
        ws.cell(row=1, column=count + 5, value="theta_rot")

        for i in range(len(self.nameList)):
            ws.cell(row=i + 2, column=1, value=self.nameList[i])
            count = 0
            for j in range(len(self.summaryIndex[i])):
                ws.cell(row=i + 2, column=count + 2,
                        value="=\'" + self.nameList[i] + "\'!L" + str(count + 2))
                count += 1
            ws.cell(row=i + 2, column=count + 2, value="=\'" + self.nameList[i] + "\'!H4")
            ws.cell(row=i + 2, column=count + 3, value="=\'" + self.nameList[i] + "\'!H5")
            ws.cell(row=i + 2, column=count + 4, value="=H" + str(i + 2) + "-0.005")
            ws.cell(row=i + 2, column=count + 5, value=0)

        # すべてのシートのデータを乗せたグラフ作成
        cht = openpyxl.chart.ScatterChart()
        cp = openpyxl.drawing.text.CharacterProperties(sz=1000)
        for i in range(len(self.nameList)):
            cht.series.append(self.srslist[i])

        cht.width = 20
        cht.height = 15
        cht.layout = openpyxl.chart.layout.Layout(manualLayout=
        openpyxl.chart.layout.ManualLayout(
            layoutTarget='inner',
            h=0.8, w=0.6,
            xMode="edge", yMode="edge"))
        self.wb['summary'].add_chart(cht, 'B15')

    def save_excel(self, excelfilename):
        self.wb.save(excelfilename)


if __name__ == '__main__':
    from pathlib import Path
    import tkinter, tkinter.filedialog, tkinter.messagebox

    root = tkinter.Tk()  # GUIのウィンドウを起動する
    root.title(u"3次元測定機生データ→excelファイル")
    root.geometry("400x100")
    root.attributes("-topmost", True)
    root.withdraw()
    tkinter.messagebox.showinfo("test", "3次元測定器の生データを選択してください(複数可)\n同名のエクセルファイルを同フォルダに出力します")
    filelist = tkinter.filedialog.askopenfilenames()  # 複数ファイル選択のダイアログ表示

    if filelist:  # ファイルリストがemptyでなければ実施
        fol = Path(filelist[0]).parent  # ファイルが保存されているフォルダ名
        con = Converter()
        for filename in filelist:
            filename = Path(filename)  # Path形式に変換
            commonname = filename.stem  # 拡張子抜きのファイル名
            inputfilename = str(filename)  # 入力ファイル
            excelfilename = str(filename.with_suffix('.xlsx'))
            print(inputfilename)  # ファイル名確認
            print(excelfilename)  # ファイル名確認
            print("\n")  # 改行
            try:
                con.convert(inputfilename, excelfilename)
            except PermissionError as e:
                print(e)
                tkinter.messagebox.showinfo("error", str(e))

    else:
        print('ファイル選択をキャンセル')

    root.quit()  # GUIのウインドウを閉じる
    root.destroy()  # GUIのウインドウを閉じる
